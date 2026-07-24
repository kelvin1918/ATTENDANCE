"""
excel_generator.py — Attendance duration-comparison report (.xlsx)
=====================================================================
Generates the duration-comparison export discussed with the instructor:
same attendance records the PDF sheet uses, plus a per-student duration
bar using Excel's own native data-bar conditional formatting — no image
rendering needed, opens natively in Excel / Google Sheets / LibreOffice.

Status colors mirror the app's own palette (local.html / admin.html) so
the exported sheet reads as the same system, not a separate design.
"""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter


def _safe(s):
    return re.sub(r'[:*?"<>|,\s]', '_', str(s)).strip('_')


def _time_to_mins(t):
    """'7:00 AM' -> 420. Returns -1 if unparseable."""
    try:
        time_part, ampm = t.strip().rsplit(' ', 1)
        h, m = (int(x) for x in time_part.split(':'))
        ampm = ampm.upper()
        if ampm == 'PM' and h != 12:
            h += 12
        if ampm == 'AM' and h == 12:
            h = 0
        return h * 60 + m
    except Exception:
        return -1


def _class_duration_minutes(time_str):
    """'7:00 AM - 11:00 AM' -> 240. Returns 0 if unparseable."""
    if not time_str or ' - ' not in time_str:
        return 0
    start_s, end_s = time_str.split(' - ', 1)
    start, end = _time_to_mins(start_s), _time_to_mins(end_s)
    if start < 0 or end < 0 or end <= start:
        return 0
    return end - start


def _fmt_duration(total_seconds):
    """Seconds -> 'Xm Ys' — matches how instructors actually read a stopwatch,
    not a decimal fraction of a minute."""
    total_seconds = round(total_seconds or 0)
    m, s = divmod(total_seconds, 60)
    return f"{m}m {s}s"


RED = "D32F2F"
STATUS_FILL = {"Present": "E8F5E9", "Late": "FFF8E1", "Partial": "FFF8E1",
                "Excused": "EEECFB", "Absent": "F3F4F6"}
STATUS_FONT = {"Present": "2E7D32", "Late": "F57F17", "Partial": "F57F17",
                "Excused": "5B4FC4", "Absent": "6B7280"}


def generate_duration_excel(class_id, subject, section, room, date,
                             time_str="", faculty_name="Instructor",
                             records=None, session_time=""):
    """
    records: list of dicts — name, sr_code, status, presence_duration_sec, note
    Returns (BytesIO buffer, filename).
    """
    if records is None:
        records = []

    filename = f"Duration_{_safe(date)}_{_safe(session_time or 'session')}_{_safe(section)}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Duration"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"{subject} — {section}"
    ws["A1"].font = Font(size=14, bold=True, color=RED)

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Room: {room}   |   Date: {date}   |   Time: {time_str}   |   Faculty: {faculty_name}"
    ws["A2"].font = Font(size=10, color="6B7280")

    header_row = 4
    headers = ["Student", "SR Code", "Status", "Time Present",
               "Class Duration", "% Attended", "Note"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=RED)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Reference "class duration" — the real elapsed camera-session length
    # (Start Camera click to Save Attendance click), stored per-record at
    # save time. Instructors don't always run a session for the exact
    # scheduled block, so this must reflect what actually happened, not the
    # schedule. Falls back to parsing the scheduled time slot only for
    # older records saved before this was tracked, and to the longest
    # on-camera time observed as a last resort.
    class_duration_min = max(
        [r.get("class_duration_min", 0) or 0 for r in records], default=0
    )
    if not class_duration_min:
        class_duration_min = _class_duration_minutes(time_str)
    if not class_duration_min:
        non_excused = [r for r in records if r.get("status") != "Excused"]
        max_dur_sec = max([r.get("presence_duration_sec", 0) or 0 for r in non_excused], default=0)
        class_duration_min = round(max_dur_sec / 60, 1) if max_dur_sec else 0

    class_duration_sec = class_duration_min * 60

    thin   = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = header_row + 1
    for r in records:
        status  = r.get("status", "Absent")
        dur_sec = r.get("presence_duration_sec", 0) or 0
        excused = status == "Excused"
        pct     = None if (excused or class_duration_sec == 0) else round((dur_sec / class_duration_sec) * 100, 1)

        values = [
            r.get("name", ""),
            r.get("sr_code", ""),
            status,
            "—" if excused else _fmt_duration(dur_sec),
            "—" if excused else _fmt_duration(class_duration_sec),
            "—" if pct is None else pct,
            r.get("note", "") or "",
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if col in (1, 7) else "center", vertical="center")
            if col == 3:
                cell.font = Font(bold=True, color=STATUS_FONT.get(status, "111827"))
                cell.fill = PatternFill("solid", fgColor=STATUS_FILL.get(status, "FFFFFF"))
        row += 1

    last_row = row - 1

    # ── Native Excel data bar on the % Attended column ──────────────────
    if last_row >= header_row + 1:
        rule = DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=100,
            color="2E7D32", showValue=True,
        )
        ws.conditional_formatting.add(f"F{header_row + 1}:F{last_row}", rule)

    widths = [26, 12, 12, 16, 16, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{header_row + 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, filename
